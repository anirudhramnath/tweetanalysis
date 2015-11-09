
import nltk
from dataclean import clean_tweets
from dataclean import clean_test_tweets
from pprint import pprint
import openpyxl

class_value_mapping={
1:"positive",
-1:"negative",
0:"neutral"
}

def get_words_in_tweets(tweets):
    all_words = []
    for (words, sentiment) in tweets:
        all_words.extend(words)
    return all_words


def get_word_features(wordlist):
    wordlist = nltk.FreqDist(wordlist)
    word_features = wordlist.keys()
    return word_features


def extract_features(document):
    document_words = set(document)
    features = {}
    for word in word_features:
        features['contains(%s)' % word] = (word in document_words)
    return features

pos_tweets = [('I love this car', 'positive'),
              ('This view is amazing', 'positive'),
              ('I feel great this morning', 'positive'),
              ('I am so excited about the concert', 'positive'),
              ('He is my best friend', 'positive')]

neg_tweets = [('I do not like this car', 'negative'),
              ('This view is horrible', 'negative'),
              ('I feel tired this morning', 'negative'),
              ('I am not looking forward to the concert', 'negative'),
              ('He is my enemy', 'negative')]
training_data_workbook = openpyxl.load_workbook("training-Obama-Romney-tweets.xlsx")
sheet_obama = training_data_workbook.get_sheet_by_name("Obama")
tweets = []
for i in range(3,7000):
    try:
        tweet_string = str(sheet_obama.cell(row=i,column=4).value)
        tweet_class = class_value_mapping[int(sheet_obama.cell(row=i,column=5).value)]
        tweets.append((tweet_string, tweet_class))
    except Exception:
        pass
    
tweets = clean_tweets(tweets)

word_features = get_word_features(get_words_in_tweets(tweets))

training_set = nltk.classify.apply_features(extract_features, tweets)
#pprint(word_features)
classifier = nltk.NaiveBayesClassifier.train(training_set)
test_tweet_list = []
for i in range(7001,7201):
    try:
        tweet_string = str(sheet_obama.cell(row=i,column=4).value)
        tweet_class = class_value_mapping[int(sheet_obama.cell(row=i,column=5).value)]
        tweet_string = clean_test_tweets([tweet_string])[0]
        test_tweet_list.append((tweet_string, tweet_class))
    except Exception:
        pass


#clean_tweet = clean_test_tweets([tweet])

for (tweet, class_type) in test_tweet_list:
  print tweet, class_type, classifier.classify(extract_features(x))
