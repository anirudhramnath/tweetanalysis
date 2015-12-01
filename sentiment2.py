
from sklearn.feature_extraction.text import TfidfTransformer
from sklearn.feature_selection import SelectKBest, chi2
from sklearn.naive_bayes import MultinomialNB
from sklearn.pipeline import Pipeline

from sklearn.svm import LinearSVC
from nltk.classify.scikitlearn import SklearnClassifier


import nltk
from dataclean import clean_tweets
from dataclean import clean_test_tweets
from pprint import pprint
import openpyxl

class_value_mapping={
1:"positive",
-1:"negative",
0:"neutral"
}

def get_features_from_tweets(tweets):
    features = []
    for (words, sentiment) in tweets:
        features.extend(words)
    return features


def get_unique_features(feature_list):
    feature_frequency_distribution = nltk.FreqDist(feature_list)
    unique_features = feature_frequency_distribution.keys()
    return unique_features


def get_feature_mapping(document):
    document_words = set(document)
    features = {}
    for word in word_features:
        features['contains(%s)' % word] = (word in document_words)
    return features

pos_tweets = [('I love this car', 'positive'),
              ('This view is amazing', 'positive'),
              ('I feel great this morning', 'positive'),
              ('I am so excited about the concert', 'positive'),
              ('He is my best friend', 'positive')]

neg_tweets = [('I do not like this car', 'negative'),
              ('This view is horrible', 'negative'),
              ('I feel tired this morning', 'negative'),
              ('I am not looking forward to the concert', 'negative'),
              ('He is my enemy', 'negative')]

training_data_workbook = openpyxl.load_workbook("training-Obama-Romney-tweets.xlsx")
sheet_obama = training_data_workbook.get_sheet_by_name("Obama")
accuracy_list = []
pos_precision_list = []
neg_precision_list = []
neu_precision_list = []
pos_recall_list = []
neg_recall_list = []
neu_recall_list = []

for fold in range(10):
    tweets = []
    list_of_input = []
    for index in range(10):
        if index != fold:
            list_of_input += range(720*index,720*index+720)
    for i in list_of_input:
        try:
            tweet_string = str(sheet_obama.cell(row=i,column=4).value)
            tweet_class = class_value_mapping[int(sheet_obama.cell(row=i,column=5).value)]
            tweets.append((tweet_string, tweet_class))
        except Exception:
            pass
    print len(tweets)
    tweets = clean_tweets(tweets)
    print len(tweets)
    word_features = get_unique_features(get_features_from_tweets(tweets))

    training_set = nltk.classify.apply_features(get_feature_mapping, tweets)
    #pprint(word_features)
    pipeline = Pipeline([('tfidf', TfidfTransformer()),
                         ('chi2', SelectKBest(chi2, k=1500)),
                         ('nb', MultinomialNB())])

    classifier = SklearnClassifier(pipeline).train(training_set)
    #classifier = SklearnClassifier(LinearSVC()).train(training_set)
    #classifier = nltk.NaiveBayesClassifier.train(training_set)
    list_of_input = range(720*fold,720*fold+720)
    test_tweet_list = []
    for i in list_of_input:
        try:
            original_tweet_string = str(sheet_obama.cell(row=i,column=4).value)
            tweet_class = class_value_mapping[int(sheet_obama.cell(row=i,column=5).value)]
            tweet_string = clean_test_tweets([original_tweet_string])[0]
            test_tweet_list.append((tweet_string, tweet_class, original_tweet_string))
        except Exception:
            pass
    
    correct, total = 0., 0.
    no_class_pos = 0.
    no_class_neg = 0.
    no_class_neu = 0.
    no_result_pos = 0.
    no_result_neg = 0.
    no_result_neu = 0.
    true_pos = 0.
    true_neg = 0.
    true_neu = 0.
    correct_class_pos = 0.
    correct_class_neg = 0.
    correct_class_neu = 0.
    
    for (tweet, class_type, original_tweet_string) in test_tweet_list:
        result = classifier.classify(extract_features(tweet))
        total += 1
        if class_type == result: correct += 1
        if class_type == 'positive':
            no_class_pos += 1
            if result == 'positive':
                correct_class_pos += 1;
        if result == 'positive':
            no_result_pos += 1
            if class_type == 'positive':
                true_pos += 1
        if class_type == 'negative':
            no_class_neg += 1
            if result == 'negative':
                correct_class_neg += 1;
        if result == 'negative':
            no_result_neg += 1
            if class_type == 'negative':
                true_neg += 1
        if class_type == 'neutral':
            no_class_neu += 1
            if result == 'neutral':
                correct_class_neu += 1;
        if result == 'neutral':
            no_result_neu += 1
            if class_type == 'neutral':
                true_neu += 1
    print 'For fold'+str(fold)
    print 'accuracy: '+str(correct * 100/total)
    accuracy_list.append(correct * 100/total)
    print 'Precision for pos:' + str(true_pos*100/no_result_pos)
    pos_precision_list.append(true_pos*100/no_result_pos)
    print 'Precision for neg:' + str(true_neg*100/no_result_neg)
    neg_precision_list.append(true_neg*100/no_result_neg)
    print 'Precision for neu:' + str(true_pos*100/no_result_neu)
    neu_precision_list.append(true_neu*100/no_result_neu)
    print 'Recall for pos:' + str(correct_class_pos*100/no_class_pos)
    pos_recall_list.append(correct_class_pos*100/no_class_pos)
    print 'Recall for neg:' + str(correct_class_neg*100/no_class_neg)
    neg_recall_list.append(correct_class_neg*100/no_class_neg)
    print 'Recall for neu:' + str(correct_class_neu*100/no_class_neu)
    neu_recall_list.append(correct_class_neu*100/no_class_neu)
    print '\n'
print "overall\n"
print 'accuracy: '+str(sum(accuracy_list)/len(accuracy_list))
print 'Precision for pos:' + str(sum(pos_precision_list)/len(pos_precision_list))
print 'Precision for neg:' + str(sum(neg_precision_list)/len(neg_precision_list))
print 'Precision for neu:' + str(sum(neu_precision_list)/len(neu_precision_list))
print 'Recall for pos:' + str(sum(pos_recall_list)/len(pos_recall_list))
print 'Recall for neg:' + str(sum(neg_recall_list)/len(neg_recall_list))
print 'Recall for neu:' + str(sum(neu_recall_list)/len(neu_recall_list))
